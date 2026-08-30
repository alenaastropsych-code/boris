"""
Собирает итоговый Excel-файл (лист PnL по неделям + сырые данные) прямо из базы,
и текстовую сводку по одной неделе для сообщения в Telegram.

Формат листа PnL — тот, что был утверждён с пользователем:
  OZON: Доходы -> Итого Ozon (нетто) -> Себестоимость -> ДОХОД OZON ЗА НЕДЕЛЮ
  WILDBERRIES: (справочно: заказы до комиссии/% комиссии) -> Доходы ->
               Итого ВБ (нетто) -> Себестоимость -> ДОХОД ВБ ЗА НЕДЕЛЮ
  ИТОГО ДОХОД ЗА НЕДЕЛЮ (Ozon + ВБ)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import db

FONT = "Arial"
MONEY_FMT = '#,##0;(#,##0);-'
PCT_FMT = '0.0%'

OZON_INCOME_GROUPS = ["Продажи", "Возвраты", "Компенсации и декомпенсации"]
OZON_EXPENSE_GROUPS = ["Вознаграждение Ozon", "Услуги доставки", "Услуги партнёров",
                       "Услуги FBO", "Продвижение и реклама", "Другие услуги и штрафы"]
WB_INCOME_GROUPS = ["Продажи", "Программа лояльности", "Возмещения логистики", "Возмещения ПВЗ"]
WB_EXPENSE_GROUPS = ["Логистика", "Хранение", "Обработка товара", "Удержания"]


def _fmt_money(v):
    return f"{v:,.0f}".replace(",", " ")


def build_week_summary_text(week_start_date) -> str:
    data = db.get_week_summary(week_start_date)
    by_group = data["by_group"]
    gross = data["gross"]
    cogs = data["cogs"]

    def g(platform, name):
        return by_group.get((platform, name), 0.0)

    oz_income = sum(g("ozon", n) for n in OZON_INCOME_GROUPS)
    oz_expense = sum(g("ozon", n) for n in OZON_EXPENSE_GROUPS)
    oz_net = oz_income + oz_expense
    oz_cogs = cogs.get("ozon", 0.0) or 0.0
    oz_profit = oz_net - oz_cogs

    wb_income = sum(g("wb", n) for n in WB_INCOME_GROUPS)
    wb_expense = sum(g("wb", n) for n in WB_EXPENSE_GROUPS)
    wb_net = wb_income + wb_expense
    wb_cogs = cogs.get("wb", 0.0) or 0.0
    wb_profit = wb_net - wb_cogs
    wb_gross = gross.get("wb", 0.0) or 0.0
    wb_commission = g("wb", "Продажи") - wb_gross
    wb_commission_pct = (-wb_commission / wb_gross * 100) if wb_gross else 0.0

    week_end = week_start_date.replace(day=week_start_date.day) + __import__("datetime").timedelta(days=6)

    lines = [
        f"📊 PnL за неделю: {week_start_date.strftime('%d.%m')}–{week_end.strftime('%d.%m.%Y')}",
        "",
    ]
    if oz_income or oz_expense:
        lines += [
            "🟦 OZON",
            f"Доходы: {_fmt_money(oz_income)} ₽",
            f"Расходы: {_fmt_money(oz_expense)} ₽",
            f"Итого Ozon (нетто): {_fmt_money(oz_net)} ₽",
            f"Себестоимость: −{_fmt_money(oz_cogs)} ₽",
            f"➜ Доход Ozon за неделю: {_fmt_money(oz_profit)} ₽",
            "",
        ]
    if wb_income or wb_expense:
        lines += [
            "🟥 WILDBERRIES",
            f"Заказов до комиссии: {_fmt_money(wb_gross)} ₽ (комиссия ВБ {wb_commission_pct:.1f}%)",
            f"Итого ВБ (нетто): {_fmt_money(wb_net)} ₽",
            f"Себестоимость: −{_fmt_money(wb_cogs)} ₽",
            f"➜ Доход ВБ за неделю: {_fmt_money(wb_profit)} ₽",
            "",
        ]
    lines += [
        "━━━━━━━━━━━━━━━",
        f"💰 ИТОГО ДОХОД ЗА НЕДЕЛЮ: {_fmt_money(oz_profit + wb_profit)} ₽",
    ]

    missing = db.get_missing_cost_articles()
    if missing:
        lines += ["", f"⚠️ Без указанной себестоимости: {len(missing)} товар(ов) — /cost чтобы добавить"]

    return "\n".join(lines)


def build_full_workbook(output_path: str):
    weeks = db.get_all_week_starts()
    if not weeks:
        raise ValueError("В базе пока нет данных")

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "PnL"

    title_font = Font(name=FONT, size=14, bold=True)
    h1_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    h1_fill = PatternFill("solid", fgColor="305496")
    section_font = Font(name=FONT, size=11, bold=True)
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    normal_font = Font(name=FONT, size=10)
    ref_font = Font(name=FONT, size=10, italic=True, color="808080")
    total_font = Font(name=FONT, size=10, bold=True)
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    net_fill_oz = PatternFill("solid", fgColor="DDEBF7")
    net_fill_wb = PatternFill("solid", fgColor="FCE4D6")
    cogs_font = Font(name=FONT, size=10, italic=True)
    income_font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    income_fill_oz = PatternFill("solid", fgColor="2E75B6")
    income_fill_wb = PatternFill("solid", fgColor="C00000")
    grand_font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
    grand_fill = PatternFill("solid", fgColor="548235")
    note_font = Font(name=FONT, size=9, italic=True, color="808080")
    center = Alignment(horizontal="center")

    total_col = 2 + len(weeks)
    ws.column_dimensions['A'].width = 48
    for i in range(len(weeks) + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 15

    ws['A1'] = "Общий PnL: Ozon + Wildberries (понедельно)"
    ws['A1'].font = title_font
    ws['A2'] = f"Собрано автоматически ботом Борис · недель в базе: {len(weeks)}"
    ws['A2'].font = Font(name=FONT, size=11, italic=True)

    header_row = 4
    ws.cell(row=header_row, column=1, value="Статья").font = h1_font
    ws.cell(row=header_row, column=1).fill = h1_fill
    for i, wstart in enumerate(weeks):
        wend = wstart + __import__("datetime").timedelta(days=6)
        label = f"{wstart.strftime('%d.%m')}–{wend.strftime('%d.%m')}"
        c = ws.cell(row=header_row, column=2 + i, value=label)
        c.font = h1_font
        c.fill = h1_fill
        c.alignment = center
    c = ws.cell(row=header_row, column=total_col, value="ИТОГО")
    c.font = h1_font
    c.fill = h1_fill
    c.alignment = center

    # собираем данные по каждой неделе одним проходом
    week_data = {w: db.get_week_summary(w) for w in weeks}

    def series_for(getter):
        vals = [getter(week_data[w]) for w in weeks]
        return vals + [sum(vals)]

    def write_line(row, label, vals, font=normal_font, indent="   "):
        ws.cell(row=row, column=1, value=indent + label).font = font
        for i, v in enumerate(vals):
            cc = ws.cell(row=row, column=2 + i, value=v)
            cc.font = font
            cc.number_format = MONEY_FMT
        return row + 1

    def section(row, title):
        ws.cell(row=row, column=1, value=title).font = section_font
        for col in range(1, total_col + 1):
            ws.cell(row=row, column=col).fill = section_fill
        return row + 1

    def platform_row(row, title, font=Font(name=FONT, size=10, bold=True, italic=True, color="595959")):
        ws.cell(row=row, column=1, value=title).font = font
        return row + 1

    row = header_row + 1

    # ================= OZON =================
    row = section(row, "OZON")
    row = platform_row(row, "Доходы")
    oz_group_vals = {}
    for name in OZON_INCOME_GROUPS + OZON_EXPENSE_GROUPS:
        vals = series_for(lambda d, n=name: d["by_group"].get(("ozon", n), 0.0))
        oz_group_vals[name] = vals
        row = write_line(row, name, vals)
        if name == OZON_INCOME_GROUPS[-1]:
            income_total = [sum(oz_group_vals[n][i] for n in OZON_INCOME_GROUPS) for i in range(len(weeks) + 1)]
            row = write_line(row, "Итого доходы Ozon", income_total, font=total_font)
            ws.cell(row=row - 1, column=1).fill = total_fill
            for col in range(2, total_col + 1):
                ws.cell(row=row - 1, column=col).fill = total_fill
            row = platform_row(row, "Расходы")
    expense_total = [sum(oz_group_vals[n][i] for n in OZON_EXPENSE_GROUPS) for i in range(len(weeks) + 1)]
    row = write_line(row, "Итого расходы Ozon", expense_total, font=total_font)
    for col in range(1, total_col + 1):
        ws.cell(row=row - 1, column=col).fill = total_fill

    oz_net = [income_total[i] + expense_total[i] for i in range(len(weeks) + 1)]
    row = write_line(row, "ИТОГО Ozon (нетто)", oz_net, font=total_font, indent="")
    for col in range(1, total_col + 1):
        ws.cell(row=row - 1, column=col).fill = net_fill_oz

    oz_cogs = [-((week_data[w]["cogs"].get("ozon") or 0.0) if w in weeks else 0.0) for w in weeks]
    oz_cogs.append(sum(oz_cogs))
    row = write_line(row, "Себестоимость товаров Ozon", oz_cogs, font=cogs_font)

    oz_profit = [oz_net[i] + oz_cogs[i] for i in range(len(weeks) + 1)]
    row = write_line(row, "ДОХОД OZON ЗА НЕДЕЛЮ", oz_profit, font=income_font, indent="")
    for col in range(1, total_col + 1):
        ws.cell(row=row - 1, column=col).fill = income_fill_oz
    row += 1

    # ================= WILDBERRIES =================
    row = section(row, "WILDBERRIES")
    gross_vals = series_for(lambda d: d["gross"].get("wb", 0.0) or 0.0)
    row = write_line(row, "Сумма заказов до вычета комиссии (справочно)", gross_vals, font=ref_font)
    sales_wb_vals = series_for(lambda d: d["by_group"].get(("wb", "Продажи"), 0.0))
    commission_vals = [sales_wb_vals[i] - gross_vals[i] for i in range(len(weeks) + 1)]
    row = write_line(row, "  из них: комиссия и эквайринг ВБ (справочно)", commission_vals, font=ref_font)
    pct_vals = [(-commission_vals[i] / gross_vals[i]) if gross_vals[i] else 0 for i in range(len(weeks) + 1)]
    ws.cell(row=row, column=1, value="  % комиссии от суммы заказов (справочно)").font = ref_font
    for i, v in enumerate(pct_vals):
        cc = ws.cell(row=row, column=2 + i, value=v)
        cc.font = ref_font
        cc.number_format = PCT_FMT
    row += 1

    row = platform_row(row, "Доходы")
    wb_group_vals = {"Продажи": sales_wb_vals}
    row = write_line(row, "Продажи (нетто, после комиссии ВБ)", sales_wb_vals)
    for name in WB_INCOME_GROUPS[1:]:
        vals = series_for(lambda d, n=name: d["by_group"].get(("wb", n), 0.0))
        wb_group_vals[name] = vals
        row = write_line(row, name, vals)
    wb_income_total = [sum(wb_group_vals[n][i] for n in WB_INCOME_GROUPS) for i in range(len(weeks) + 1)]
    row = write_line(row, "Итого доходы ВБ", wb_income_total, font=total_font)
    for col in range(1, total_col + 1):
        ws.cell(row=row - 1, column=col).fill = total_fill

    row = platform_row(row, "Расходы")
    wb_exp_vals = {}
    for name in WB_EXPENSE_GROUPS:
        vals = series_for(lambda d, n=name: d["by_group"].get(("wb", n), 0.0))
        wb_exp_vals[name] = vals
        row = write_line(row, name, vals)
    wb_expense_total = [sum(wb_exp_vals[n][i] for n in WB_EXPENSE_GROUPS) for i in range(len(weeks) + 1)]
    row = write_line(row, "Итого расходы ВБ", wb_expense_total, font=total_font)
    for col in range(1, total_col + 1):
        ws.cell(row=row - 1, column=col).fill = total_fill

    wb_net = [wb_income_total[i] + wb_expense_total[i] for i in range(len(weeks) + 1)]
    row = write_line(row, "ИТОГО ВБ (нетто)", wb_net, font=total_font, indent="")
    for col in range(1, total_col + 1):
        ws.cell(row=row - 1, column=col).fill = net_fill_wb

    wb_cogs = [-((week_data[w]["cogs"].get("wb") or 0.0)) for w in weeks]
    wb_cogs.append(sum(wb_cogs))
    row = write_line(row, "Себестоимость товаров ВБ", wb_cogs, font=cogs_font)

    wb_profit = [wb_net[i] + wb_cogs[i] for i in range(len(weeks) + 1)]
    row = write_line(row, "ДОХОД ВБ ЗА НЕДЕЛЮ", wb_profit, font=income_font, indent="")
    for col in range(1, total_col + 1):
        ws.cell(row=row - 1, column=col).fill = income_fill_wb
    row += 2

    grand = [oz_profit[i] + wb_profit[i] for i in range(len(weeks) + 1)]
    row = write_line(row, "ИТОГО ДОХОД ЗА НЕДЕЛЮ (Ozon + ВБ)", grand, font=grand_font, indent="")
    for col in range(1, total_col + 1):
        ws.cell(row=row - 1, column=col).fill = grand_fill
    row += 1

    for note in [
        "⚠ «Доход за неделю» — после себестоимости, но ещё без налогов, ФОТ, аренды и прочих расходов бизнеса.",
        "Себестоимость считается по таблице product_costs — умножается на реально проданное количество.",
        "Строки «Сумма заказов до вычета комиссии», «комиссия и эквайринг ВБ» и «% комиссии» — справочные.",
        "Недели считаются Пн–Вс.",
    ]:
        ws.cell(row=row, column=1, value=note).font = note_font
        row += 1

    ws.freeze_panes = "B6"
    wb_out.save(output_path)
    return output_path
